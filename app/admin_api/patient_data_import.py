from werkzeug.datastructures import FileStorage
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from web_errors import WebError
from dataclasses import dataclass
from typing import Iterable
from clinics.data_access import get_most_common_clinic
from patients.data_access import patient_from_key_data, add_patient
from patients.patient import Patient
from visits.data_access import first_visit_by_patient_and_date, add_visit
from visits.visit import Visit
from events.data_access import clear_all_events, add_event
from events.event import Event
import uuid
from language_strings.language_string import LanguageString
from util import as_string
from datetime import date, timedelta, datetime
import itertools
import json
from config import DEFAULT_PROVIDER_ID_FOR_IMPORT
import pandas as pd
import dateutil

COLUMNS = ['visit_date',
           'first_name',
           'surname',
           'age',
           'gender',
           'home_country',
           'phone',
           # patient details
           'medical_num',
           'dental_num',
           'optometry_num',
           'patient_id',
           'community',
           'zone',
           'block',
           'lot',
           'emergency_phone',
           'mother',
           'mother_phone',
           'father',
           'father_phone',
           'partner',
           'partner_phone',
           'employer',
           'insurance',
           # vitals
           'doctor_v',
           'glycemia',
           'weight',
           'weight_lb',
           'ideal_weight',
           'blood_pressure',
           'pulse',
           'respiration',
           'o2_sats',
           'height',
           'temperature',
           'blood_type',
           'notes',
           # medical_hx
           'doctor_mh',
           'malnutrition',
           'prenatal',
           'sexual_hx',
           'nutrition',
           'parasite_treatment',
           'family_hx',
           'surgery_hx',
           'vaccinations',
           # evaluation
           'doctor_ce',
           'visit_date',
           'reason',
           'observations',
           'medications',
           'breast_exam',
           'diagnosis',
           'treatment',
           'community_visit',
           'promoter_visit',
           'refusal',
           'next_visit_date',
           'next_visit_reason',
           # Med from stock
           'doctor_s',
           'medicine_s',
           'format_s',
           'dosage_s',
           'days_s',
           # Med OTC
           'doctor_otc',
           'medicine_otc',
           'format_otc',
           'dosage_otc',
           'days_otc',
           # Controlled med
           'doctor_c',
           'medicine_c',
           'format_c',
           'dosage_c',
           'days_c',
           # Med pathologies
           'doctor_mp',
           'miscarriages',
           'food_allergies',
           'animal_allergies',
           'atmosphere_allergies',
           'insect_allergies',
           'latex_allergies',
           'medicine_allergies',
           'other_allergies',
           'tonsillitis',
           'anemic',
           'arthritis',
           'asthma',
           'neck_pain',
           'cervicovaginitis',
           'c_section',
           'sciatic_pain',
           'cholesterol',
           'infant_colic',
           'conjunctivitis',
           'covid',
           'malnourishment',
           'diabetes',
           'migraines',
           'diarrhea',
           'ecocardiogram',
           'electrocardiogram',
           'pregnant',
           'pregnancies',
           'chikungunya',
           'dengue',
           'malaria',
           'other_mosquito',
           'zika',
           'copd',
           'gastritis',
           'scabies',
           'last_PAP',
           'vaginal_fluid',
           'hypertension',
           'hypothyroidism',
           'bacterial_resp',
           'viral_resp',
           'uti',
           'renal_failure',
           'breastfeeding',
           'lumbago',
           'menopause',
           'nausea',
           'nephrolithiasis_renal',
           'diabetic_neuropathy',
           'obesity',
           'osteoarthritis',
           'otitis',
           'paralysis',
           'parasites',
           'skin_healthy',
           'skin_ulcers',
           'skin_infected',
           'lice',
           'postnatal_visit',
           'prenatal_visit',
           'eye_prob',
           'emotional_prob',
           'gynecological_prob',
           'parkinsons',
           'epilepsy',
           'neurological_prob',
           'therapist_referred',
           'developmentally_delayed',
           'vitamins',
           'last_menstruation',
           'hiv',
           'vomiting',
           'other_mp',
           # Psych pathologies
           'doctor_pp',
           'anxiety',
           'nuclear_family',
           'self_esteem',
           'attention_deficit',
           'depression',
           'grief',
           'stress',
           'disfunctional_family',
           'hyperactivity',
           'inappropriate_play',
           'language_problems',
           'behavioral_problems',
           'school_problems',
           'psychosis',
           'suicidal',
           'personality_disorders',
           'trauma',
           'psychological_evaluations',
           'domestic_violence_family',
           'domestic_violence_spouse',
           'referral_hospital',
           'other_pp',
           # Household env
           'doctor_he',
           'potable_water',
           'animals',
           'gas_cooking',
           'wood_cooking',
           'household_size',
           'toilet',
           'latrine',
           'family_violence',
           # Lab orders
           'doctor_lo',
           'hematic_biometry',
           'urinalysis',
           'biochemistry',
           'lipid_profile',
           'pregnancy_test',
           'immunology_test',
           'PAP_test',
           'serology_test',
           'stool_test',
           'fecal_antigens',
           'blood_type_lo',
           'HIV_test',
           'other_lo',
           # Lab tests
           'doctor_lt',
           'CBC_Immature',
           'CBCMCHC',
           'CBCMCH',
           'CBC_basophils',
           'CBC_eosinophils',
           'CBC_hematocrit',
           'CBC_hemoglobin',
           'CBC_lymphocytes',
           'CBC_monocytes',
           'CBC_platelets',
           'CBC_segmented',
           'CBCWBC',
           'CBC_platelet_count',
           'CBCRBC',
           'CBCMCV',
           'biochem_uric',
           'biochem_creatinine',
           'biochem_glucose',
           'lipid_cholesterol',
           'lipid_HDL',
           'lipid_LDL',
           'lipid_triglycerides',
           'lipid_VLDL',
           'fecal_antigens',
           'fecal_mononuclear',
           'fecal_polymophonuclear',
           'fecal_bacteria',
           'fecal_erythrocytes',
           'fecal_leukocytes',
           'fecal_others',
           'fecal_cysts',
           'fecal_trophozoites',
           'fecal_PH',
           'fecal_reducers',
           'fecal_occult',
           'fecal_color',
           'fecal_observations',
           'fecal_consistency',
           'pregnancy_hemogravindex',
           'pregnancy_gravindex',
           'microbiology_pilori',
           'microbiology_malaria',
           'serology_strep',
           'serology_rheumatoid',
           'serology_others',
           'serology_protein',
           'serology_VDRL',
           'serology_VIH',
           'serology_VSG',
           'blood_type_lt',
           'RH_factor',
           # Lab urine tests
           'doctor_lu',
           'color_physical',
           'aspects_physical',
           'sediment_physical',
           'density_physical',
           'proteins_chem',
           'hemoglobin_chem',
           'ketonic_chem',
           'pH_chem',
           'urobilinogen_chem',
           'glucose_chem',
           'bilirubins_chem',
           'leukocytes_chem',
           'nitrite_chem',
           'epithelial_micro',
           'leukocytes_micro',
           'erythrocytes_micro',
           'cylinders_micro',
           'crystals_micro',
           'bacteria_micro',
           'yeasts_micro',
           'cRenal_micro',
           'h_mucous_micro',
           'observations_micro',
           # PAP
           'doctor_pap',
           'adequate_cyto',
           'incomplete_cyto',
           'cellularity_cyto',
           'fixation_cyto',
           'hemorrhage_cyto',
           'exudate_cyto',
           'endocervical_cyto',
           'inadequate_cyto',
           'bleeding_cyto',
           'fixation_inadequate_cyto',
           'squamous',
           'US_squamous',
           'H_squamous',
           'lgsil',
           'cellular_lgsil',
           'dysplasia_lgsil',
           'hgsil',
           'cin_hgsil',
           'dysplasia_hgsil',
           'carcinoma_hgsil',
           'neg_intraepithelial',
           'ais_carcinoma',
           'invasive_carcinoma',
           'agc_carcinoma',
           'endocervical_carcinoma',
           'endometrial_carcinoma',
           'neoplasia_carcinoma',
           'nos_carcinoma',
           'iscc',
           'atrophy',
           'coccoid',
           'regenerative',
           'candida',
           'bv',
           'herpes',
           'inflammation',
           'trichomonas',
           'other',
           # Ultrasound
           'doctor_u',
           'abdominal',
           'joints',
           'neck',
           'doppler',
           'breast',
           'first_OB',
           'second_OB',
           'third_OB',
           'soft_tissues',
           'pelvic',
           'prostate',
           'renal',
           'testicular',
           'other',
           'goiter',
           'wrapped_cord',
           'cholelithiasis',
           'prostate1',
           'prostate2',
           'prostate3',
           'prostate4',
           'endometrial_thickening',
           'splenomegaly',
           'mild_hepatic',
           'moderate_hepatic',
           'severe_hepatic',
           'hepatomegaly',
           'cirrhosis',
           'venous_insufficiency',
           'mild_renal',
           'moderate_renal',
           'severe_renal',
           'nephrolithiasis',
           'complex_masses',
           'benign_nodules',
           'malignant_nodules',
           'thyroid_nodules',
           'normal_ultrasound',
           'polycystic_ovaries',
           'simple_ovarian_cysts',
           # Family Planning
           'doctor_fp',
           'age_sexual_relations',
           'number_children',
           'more_children',
           'abnormal_bleeding',
           'varicose',
           'tubal_sterilization',
           'vasectomy_sterilization',
           'std_type',
           'injection',
           'iud_BC',
           'implant_BC',
           'pills_BC',
           'condoms',
           'sex_orientation',
           'married',
           'permanent_partner',
           # Dental Origin
           'doctor_do',
           'bernabe',
           'canon',
           'club_la_esperanza',
           'el_shaddai',
           'hogar',
           'jinotepe',
           'la_hermosa',
           'leon_de_juda',
           'los_ninn',
           'luz_vida',
           'manantial_de_vida',
           'maranatha',
           'matagalpa',
           'monte_de_sion',
           'nueva_jerusalen',
           'obrero_levitico',
           'posoltega_jose_lara',
           'rios_de_agua_viva',
           'rivas',
           'verbo_sur',
           'jubilee_house_community',
           'other_feeding_center',
           # Dental Treatment
           'doctor_dt',
           'toothbrush',
           'consult',
           'extraction',
           'fluoride',
           'floss',
           'cleaning_first',
           'cleaning_firstYear',
           'cleaning_second',
           'toothpaste',
           'xray',
           'amalgama_restoration',
           'metal_restoration',
           'ionomero_restoration',
           'mri_restoration',
           'space_restoration',
           'resin_restoration',
           'zoe_restoration',
           'acetate_restoration',
           'pulpotomy_restoration',
           'sd_fluoride',
           'sealant',
           'teaching',
           'impression',
           'other',
           # Program Trainings
           'doctor_pt',
           'asthma_subject',
           'pregnancy_subject',
           'dispensarizados1_subject',
           'dispensarizados2_subject',
           'dispensarizados3_subject',
           'dispensarizados4_subject',
           'breastfeeding_subject',
           'girls_group_subject',
           'lbgtq_subject',
           'boys_older_subject',
           'boys_younger_subject',
           'new_mothers_subject',
           'family_planning_subject',
           'toddler_mothers_subject',
           'health_promoters_subject',
           'hiv_subject',
           'other_subject',
           # Xray orders
           'doctor_xo',
           'hip_antero',
           'hip_lateral',
           'hip_posterior',
           'retrograde_cystogram',
           'cranial_antero',
           'cranial_lateral',
           'cranial_posterior',
           'femur_antero',
           'femur_lateral',
           'femur_posterior',
           'foot_antero',
           'standing_lateral',
           'standing_oblique',
           'foot_posterior',
           'iv_pyelogram',
           'knee_antero',
           'knee_lateral',
           'knee_posterior',
           'tibia_antero',
           'tibia_lateral',
           'tibia_posterior',
           'ankle_antero',
           'ankle_lateral',
           'ankle_posterior',
           'chest_antero',
           'chest_lateral',
           'chest_posterior',
           # Xray results
           'doctor_xr',
           'arthritis',
           'cardiomegaly',
           'ureteral_duplication',
           'lung_disease',
           'fibrosis',
           'fractures',
           'pneumonia',
           'pulmonary_nodule',
           'osteoarthritis_xray',
           'osteoporosis',
           'ectopic_kidney',
           'rinon_herradura',
           'tumor',
           # Optometry
           'doctor_o',
           'cataract',
           'cataract_operation',
           'glasses',
           'glasses_duration',
           'orange_reflection',
           'visualAcuity_OD',
           'visualAcuity_OS',
           'autorefractor_OD',
           'autorefractor_OS',
           'phoropter_OD',
           'phoropter_OS',
           'phoropterADD',
           'lenses_OD',
           'lenses_OS',
           'lenses_ADD',
           'visual_acuity_lenses_OD',
           'visual_acuity_lenses_OS',
           # Accident Report
           'doctor_ar',
           'accident_work',
           'description',
           'trauma',
           'trauma_details',
           'referral_specialist',
           'follow_up_visit',
           'other_ar'
           ]


@dataclass
class PatientDataRow:
    visit_date: datetime = None
    first_name: str = None
    surname: str = None
    age: str = None
    gender: str = None
    home_country: str = None
    phone: str = None
    # patient details -- need event export still
    medical_num: str = None
    dental_num: str = None
    optometry_num: str = None
    patient_id: str = None
    community: str = None
    zone: str = None
    block: str = None
    lot: str = None
    emergency_phone: str = None
    mother: str = None
    mother_phone: str = None
    father: str = None
    father_phone: str = None
    partner: str = None
    partner_phone: str = None
    employer: str = None
    insurance: str = None
    # vitals
    doctor_v: str = None
    glycemia: float = None
    weight: float = None
    weight_lb: float = None
    ideal_weight: float = None
    blood_pressure: str = None
    pulse: float = None
    respiration: float = None
    o2_sats: float = None
    height: float = None
    temperature: float = None
    blood_type: str = None
    notes: str = None
    # medical_hx
    doctor_mh: str = None
    malnutrition: str = None
    prenatal: str = None
    sexual_hx: str = None
    nutrition: str = None
    parasite_treatment: str = None
    family_hx: str = None
    surgery_hx: str = None
    vaccinations: str = None
    # evaluation
    doctor_ce: str = None
    visit_date: str = None
    reason: str = None
    observations: str = None
    medications: str = None
    breast_exam: str = None
    diagnosis: str = None
    treatment: str = None
    community_visit: str = None
    promoter_visit: str = None
    refusal: str = None
    next_visit_date: datetime = None
    next_visit_reason: str = None
    # Med from stock
    doctor_s: str = None
    medicine_s: str = None
    format_s: str = None
    dosage_s: str = None
    days_s: float = None
    # Med OTC
    doctor_otc: str = None
    medicine_otc: str = None
    format_otc: str = None
    dosage_otc: str = None
    days_otc: float = None
    # Controlled med
    doctor_c: str = None
    medicine_c: str = None
    format_c: str = None
    dosage_c: str = None
    days_c: float = None
    # Med pathologies
    doctor_mp: str = None
    miscarriages: str = None
    food_allergies: str = None
    animal_allergies: str = None
    atmosphere_allergies: str = None
    insect_allergies: str = None
    latex_allergies: str = None
    medicine_allergies: str = None
    other_allergies: str = None
    tonsillitis: str = None
    anemic: str = None
    arthritis: str = None
    asthma: str = None
    neck_pain: str = None
    cervicovaginitis: str = None
    c_section: str = None
    sciatic_pain: str = None
    cholesterol: str = None
    infant_colic: str = None
    conjunctivitis: str = None
    covid: str = None
    malnourishment: str = None
    diabetes: str = None
    migraines: str = None
    diarrhea: str = None
    ecocardiogram: str = None
    electrocardiogram: str = None
    pregnant: str = None
    pregnancies: str = None
    chikungunya: str = None
    dengue: str = None
    malaria: str = None
    other_mosquito: str = None
    zika: str = None
    copd: str = None
    gastritis: str = None
    scabies: str = None
    last_PAP: str = None
    vaginal_fluid: str = None
    hypertension: str = None
    hypothyroidism: str = None
    bacterial_resp: str = None
    viral_resp: str = None
    uti: str = None
    renal_failure: str = None
    breastfeeding: str = None
    lumbago: str = None
    menopause: str = None
    nausea: str = None
    nephrolithiasis_renal: str = None
    diabetic_neuropathy: str = None
    obesity: str = None
    osteoarthritis: str = None
    otitis: str = None
    paralysis: str = None
    parasites: str = None
    skin_healthy: str = None
    skin_ulcers: str = None
    skin_infected: str = None
    lice: str = None
    postnatal_visit: str = None
    prenatal_visit: str = None
    eye_prob: str = None
    emotional_prob: str = None
    gynecological_prob: str = None
    parkinsons: str = None
    epilepsy: str = None
    neurological_prob: str = None
    therapist_referred: str = None
    developmentally_delayed: str = None
    vitamins: str = None
    last_menstruation: str = None
    hiv: str = None
    vomiting: str = None
    other_mp: str = None
    # Psych pathologies
    doctor_pp: str = None
    anxiety: str = None
    nuclear_family: str = None
    self_esteem: str = None
    attention_deficit: str = None
    depression: str = None
    grief: str = None
    stress: str = None
    disfunctional_family: str = None
    hyperactivity: str = None
    inappropriate_play: str = None
    language_problems: str = None
    behavioral_problems: str = None
    school_problems: str = None
    psychosis: str = None
    suicidal: str = None
    personality_disorders: str = None
    trauma: str = None
    psychological_evaluations: str = None
    domestic_violence_family: str = None
    domestic_violence_spouse: str = None
    referral_hospital: str = None
    other_pp: str = None
    # Household env
    doctor_he: str = None
    potable_water: bool = None
    animals: str = None
    gas_cooking: bool = None
    wood_cooking: bool = None
    household_size: float = None
    toilet: bool = None
    latrine: bool = None
    family_violence: str = None
    # Lab orders
    doctor_lo: bool = None
    hematic_biometry: bool = None
    urinalysis: bool = None
    biochemistry: bool = None
    lipid_profile: bool = None
    pregnancy_test: bool = None
    immunology_test: bool = None
    PAP_test: bool = None
    serology_test: bool = None
    stool_test: bool = None
    fecal_antigens: bool = None
    blood_type_lo: bool = None
    HIV_test: bool = None
    other_lo: str = None
    # Lab tests
    doctor_lt: str = None
    CBC_Immature: bool = None
    CBCMCHC: bool = None
    CBCMCH: bool = None
    CBC_basophils: bool = None
    CBC_eosinophils: bool = None
    CBC_hematocrit: bool = None
    CBC_hemoglobin: bool = None
    CBC_lymphocytes: bool = None
    CBC_monocytes: bool = None
    CBC_platelets: bool = None
    CBC_segmented: bool = None
    CBCWBC: bool = None
    CBC_platelet_count: bool = None
    CBCRBC: bool = None
    CBCMCV: bool = None
    biochem_uric: bool = None
    biochem_creatinine: bool = None
    biochem_glucose: bool = None
    lipid_cholesterol: bool = None
    lipid_HDL: bool = None
    lipid_LDL: bool = None
    lipid_triglycerides: bool = None
    lipid_VLDL: bool = None
    fecal_antigens: bool = None
    fecal_mononuclear: bool = None
    fecal_polymophonuclear: bool = None
    fecal_bacteria: bool = None
    fecal_erythrocytes: bool = None
    fecal_leukocytes: bool = None
    fecal_others: bool = None
    fecal_cysts: bool = None
    fecal_trophozoites: bool = None
    fecal_PH: bool = None
    fecal_reducers: bool = None
    fecal_occult: bool = None
    fecal_color: bool = None
    fecal_observations: bool = None
    fecal_consistency: bool = None
    pregnancy_hemogravindex: bool = None
    pregnancy_gravindex: bool = None
    microbiology_pilori: bool = None
    microbiology_malaria: bool = None
    serology_strep: bool = None
    serology_rheumatoid: bool = None
    serology_others: bool = None
    serology_protein: bool = None
    serology_VDRL: bool = None
    serology_VIH: bool = None
    serology_VSG: bool = None
    blood_type_lt: bool = None
    RH_factor: bool = None
    # Lab urine tests
    doctor_lu: bool = None
    color_physical: bool = None
    aspects_physical: bool = None
    sediment_physical: bool = None
    density_physical: bool = None
    proteins_chem: bool = None
    hemoglobin_chem: bool = None
    ketonic_chem: bool = None
    pH_chem: bool = None
    urobilinogen_chem: bool = None
    glucose_chem: bool = None
    bilirubins_chem: bool = None
    leukocytes_chem: bool = None
    nitrite_chem: bool = None
    epithelial_micro: bool = None
    leukocytes_micro: bool = None
    erythrocytes_micro: bool = None
    cylinders_micro: bool = None
    crystals_micro: bool = None
    bacteria_micro: bool = None
    yeasts_micro: bool = None
    cRenal_micro: bool = None
    h_mucous_micro: bool = None
    observations_micro: bool = None
    # PAP
    doctor_pap: str = None
    adequate_cyto: str = None
    incomplete_cyto: str = None
    cellularity_cyto: str = None
    fixation_cyto: str = None
    hemorrhage_cyto: str = None
    exudate_cyto: str = None
    endocervical_cyto: str = None
    inadequate_cyto: str = None
    bleeding_cyto: str = None
    fixation_inadequate_cyto: str = None
    squamous: str = None
    US_squamous: str = None
    H_squamous: str = None
    lgsil: str = None
    cellular_lgsil: str = None
    dysplasia_lgsil: str = None
    hgsil: str = None
    cin_hgsil: str = None
    dysplasia_hgsil: str = None
    carcinoma_hgsil: str = None
    neg_intraepithelial: str = None
    ais_carcinoma: str = None
    invasive_carcinoma: str = None
    agc_carcinoma: str = None
    endocervical_carcinoma: str = None
    endometrial_carcinoma: str = None
    neoplasia_carcinoma: str = None
    nos_carcinoma: str = None
    iscc: str = None
    atrophy: str = None
    coccoid: str = None
    regenerative: str = None
    candida: str = None
    bv: str = None
    herpes: str = None
    inflammation: str = None
    trichomonas: str = None
    other: str = None
    # Ultrasound
    doctor_u: bool = None
    abdominal: bool = None
    joints: bool = None
    neck: bool = None
    doppler: bool = None
    breast: bool = None
    first_OB: bool = None
    second_OB: bool = None
    third_OB: bool = None
    soft_tissues: bool = None
    pelvic: bool = None
    prostate: bool = None
    renal: bool = None
    testicular: bool = None
    other: bool = None
    goiter: str = None
    wrapped_cord: str = None
    cholelithiasis: str = None
    prostate1: str = None
    prostate2: str = None
    prostate3: str = None
    prostate4: str = None
    endometrial_thickening: str = None
    splenomegaly: str = None
    mild_hepatic: str = None
    moderate_hepatic: str = None
    severe_hepatic: str = None
    hepatomegaly: str = None
    cirrhosis: str = None
    venous_insufficiency: str = None
    mild_renal: str = None
    moderate_renal: str = None
    severe_renal: str = None
    nephrolithiasis: str = None
    complex_masses: str = None
    benign_nodules: str = None
    malignant_nodules: str = None
    thyroid_nodules: str = None
    normal_ultrasound: str = None
    polycystic_ovaries: str = None
    simple_ovarian_cysts: str = None
    # Family Planning
    doctor_fp: str = None
    age_sexual_relations: float = None
    number_children: float = None
    more_children: bool = None
    abnormal_bleeding: bool = None
    varicose: bool = None
    tubal_sterilization: bool = None
    vasectomy_sterilization: bool = None
    std_type: bool = None
    injection: bool = None
    iud_BC: bool = None
    implant_BC: bool = None
    pills_BC: bool = None
    condoms: bool = None
    sex_orientation: bool = None
    married: bool = None
    permanent_partner: bool = None
    # Dental Origin
    doctor_do: str = None
    bernabe: str = None
    canon: str = None
    club_la_esperanza: str = None
    el_shaddai: str = None
    hogar: str = None
    jinotepe: str = None
    la_hermosa: str = None
    leon_de_juda: str = None
    los_ninn: str = None
    luz_vida: str = None
    manantial_de_vida: str = None
    maranatha: str = None
    matagalpa: str = None
    monte_de_sion: str = None
    nueva_jerusalen: str = None
    obrero_levitico: str = None
    posoltega_jose_lara: str = None
    rios_de_agua_viva: str = None
    rivas: str = None
    verbo_sur: str = None
    jubilee_house_community: str = None
    other_feeding_center: str = None
    # Dental Treatment
    doctor_dt: str = None
    toothbrush: str = None
    consult: str = None
    extraction: str = None
    fluoride: str = None
    floss: str = None
    cleaning_first: str = None
    cleaning_firstYear: str = None
    cleaning_second: str = None
    toothpaste: str = None
    xray: str = None
    amalgama_restoration: str = None
    metal_restoration: str = None
    ionomero_restoration: str = None
    mri_restoration: str = None
    space_restoration: str = None
    resin_restoration: str = None
    zoe_restoration: str = None
    acetate_restoration: str = None
    pulpotomy_restoration: str = None
    sd_fluoride: str = None
    sealant: str = None
    teaching: str = None
    impression: str = None
    other: str = None
    # Program Trainings
    doctor_pt: str = None
    asthma_subject: str = None
    pregnancy_subject: str = None
    dispensarizados1_subject: str = None
    dispensarizados2_subject: str = None
    dispensarizados3_subject: str = None
    dispensarizados4_subject: str = None
    breastfeeding_subject: str = None
    girls_group_subject: str = None
    lbgtq_subject: str = None
    boys_older_subject: str = None
    boys_younger_subject: str = None
    new_mothers_subject: str = None
    family_planning_subject: str = None
    toddler_mothers_subject: str = None
    health_promoters_subject: str = None
    hiv_subject: str = None
    other_subject: str = None
    # Xray orders
    doctor_xo: str = None
    hip_antero: bool = None
    hip_lateral: bool = None
    hip_posterior: bool = None
    retrograde_cystogram: bool = None
    cranial_antero: bool = None
    cranial_lateral: bool = None
    cranial_posterior: bool = None
    femur_antero: bool = None
    femur_lateral: bool = None
    femur_posterior: bool = None
    foot_antero: bool = None
    standing_lateral: bool = None
    standing_oblique: bool = None
    foot_posterior: bool = None
    iv_pyelogram: bool = None
    knee_antero: bool = None
    knee_lateral: bool = None
    knee_posterior: bool = None
    tibia_antero: bool = None
    tibia_lateral: bool = None
    tibia_posterior: bool = None
    ankle_antero: bool = None
    ankle_lateral: bool = None
    ankle_posterior: bool = None
    chest_antero: bool = None
    chest_lateral: bool = None
    chest_posterior: bool = None
    # Xray results
    doctor_xr: str = None
    arthritis: str = None
    cardiomegaly: str = None
    ureteral_duplication: str = None
    lung_disease: str = None
    fibrosis: str = None
    fractures: str = None
    pneumonia: str = None
    pulmonary_nodule: str = None
    osteoarthritis_xray: str = None
    osteoporosis: str = None
    ectopic_kidney: str = None
    rinon_herradura: str = None
    tumor: str = None
    # Optometry
    doctor_o: str = None
    cataract: float = None
    cataract_operation: float = None
    glasses: float = None
    glasses_duration: float = None
    orange_reflection: float = None
    visualAcuity_OD: float = None
    visualAcuity_OS: float = None
    autorefractor_OD: float = None
    autorefractor_OS: float = None
    phoropter_OD: float = None
    phoropter_OS: float = None
    phoropterADD: float = None
    lenses_OD: float = None
    lenses_OS: float = None
    lenses_ADD: float = None
    visual_acuity_lenses_OD: float = None
    visual_acuity_lenses_OS: float = None
    # Accident Report
    doctor_ar: str = None
    accident_work: bool = None
    description: bool = None
    trauma: bool = None
    trauma_details: bool = None
    referral_specialist: bool = None
    follow_up_visit: datetime = None
    other_ar: str = None


# COLUMN_TYPES = [str, None, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, float, str,
#                 float, float, float, float, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str,
#                 str, str]


# class PatientDataImporter:
#     def __init__(self, data_file: FileStorage):
#         self.data_filename = self._write_file_to_tempfile(data_file)

#     def run(self):
#         all_rows = [self._parse_row(row) for row in self.iter_data_rows()]
#         print('Creating patients...')
#         self._create_patients(all_rows)
#         print('Creating visits...')
#         self._create_visits(all_rows)

#     def _parse_row(self, row):
#         if len(row) != 41:
#             raise WebError('All data rows must have exactly 41 data points.', 400)
#         values = [self._parse_cell(value, formatter) for value, formatter in zip(row, COLUMN_TYPES)]
#         return PatientDataRow(**dict(zip(COLUMNS, values)))

#     def _parse_cell(self, cell, formatter):
#         if cell == 'Nil' or cell is None:
#             return None
#         if formatter is None:
#             return cell
#         return formatter(cell)

#     @staticmethod
#     def _write_file_to_tempfile(data_file: FileStorage):
#         handle = NamedTemporaryFile('wb', delete=False, suffix='.xlsx')
#         data_file.save(handle)
#         handle.close()
#         print('Upload written to', handle.name)
#         return handle.name

#     def iter_data_rows(self):
#         wb = load_workbook(self.data_filename)
#         ws = wb.active
#         for idx, row in enumerate(ws.iter_rows(min_row=3, max_col=41, values_only=True)):
#             if all(x is None for x in row):
#                 continue
#             yield row

#     def _create_patients(self, rows: Iterable[PatientDataRow]):
#         for patient_data in set(map(lambda r: (r.first_name, r.surname, r.gender, r.home_country, r.age), rows)):
#             first_name, surname, gender, home_country, age = patient_data
#             if not patient_from_key_data(first_name, surname, home_country, self._parse_sex(gender)):
#                 self._create_patient(first_name, surname, home_country, gender, age)

#     def _create_patient(self, given_name, surname, home_country, sex, age):
#         given_name_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': given_name})
#         surname_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': surname})
#         inferred_dob = self._infer_dob(age)
#         patient = Patient(
#             id=str(uuid.uuid4()),
#             edited_at=datetime.now(),
#             given_name=given_name_ls,
#             surname=surname_ls,
#             date_of_birth=inferred_dob,
#             sex=self._parse_sex(sex),
#             country=LanguageString(id=str(uuid.uuid4()), content_by_language={'en': home_country}),
#             phone=None,
#             hometown=None
#         )
#         add_patient(patient)

#     @staticmethod
#     def _parse_sex(sex_str):
#         if sex_str is None:
#             return None
#         elif 'm' in sex_str.lower():
#             return 'M'
#         elif 'f' in sex_str.lower():
#             return 'F'
#         else:
#             return None

#     def _infer_dob(self, age_string):
#         try:
#             int_prefix = int(''.join(itertools.takewhile(str.isnumeric, age_string)))
#             today = date.today()
#             if 'months' in age_string:
#                 return today - timedelta(days=30 * int_prefix)
#             elif 'weeks' in age_string:
#                 return today - timedelta(weeks=int_prefix)
#             elif 'days' in age_string:
#                 return today - timedelta(days=int_prefix)
#             else:
#                 # Assume years if no unit is specified
#                 return today - timedelta(days=365 * int_prefix)
#         except (ValueError, TypeError):
#             return date(1900, 1, 1)

#     @staticmethod
#     def _parse_date(date_str):
#         if isinstance(date_str, date) or isinstance(date_str, datetime):
#             return date_str
#         try:
#             dt = pd.to_datetime(date_str, dayfirst=True).to_pydatetime()
#             return date(year=dt.year, month=dt.month, day=dt.day)
#         except dateutil.parser._parser.ParserError:
#             return None

#     def _create_visits(self, rows: Iterable[PatientDataRow]):
#         for row in rows:
#             patient_id = patient_from_key_data(row.first_name, row.surname, row.home_country, self._parse_sex(row.gender))
#             if not patient_id:
#                 print('Warning: unknown patient; skipping.')
#                 continue
#             visit_date = self._parse_date(row.visit_date)
#             visit_id, visit_timestamp = first_visit_by_patient_and_date(patient_id, visit_date)

#             # TODO: The data import format does not currently specify a clinic. Since
#             # current Hikma instances are single clinic anyway, just get the most common
#             # clinic (in case there is a demo one with few if any visits) and use that.
#             clinic_id = get_most_common_clinic()

#             # TODO: The data import format does not currently specify a provider in a format
#             # that we can use. So for now, use a per-instance default provider that is set via
#             # environment variable.
#             provider_id = DEFAULT_PROVIDER_ID_FOR_IMPORT

#             if visit_id is None:
#                 visit_id = str(uuid.uuid4())
#                 visit_timestamp = datetime.combine(visit_date, datetime.min.time())
#                 visit = Visit(
#                     id=visit_id,
#                     patient_id=patient_id,
#                     edited_at=datetime.now(),
#                     clinic_id=clinic_id,
#                     provider_id=provider_id,
#                     check_in_timestamp=visit_timestamp
#                 )
#                 add_visit(visit)

#                 # Until we implement full deletion, only add visit the first time it is seen.
#                 self._update_events(patient_id, visit_id, visit_timestamp, row)

#     def _update_events(self, patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         # TODO: This will need to be replaced with a mode of deletion that persists through synchronization.
#         # clear_all_events(visit_id)
#         if row.allergies:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Allergies', row.allergies)
#         if any([row.dispensed_medicine_1, row.dispensed_medicine_2,
#                 row.dispensed_medicine_3, row.dispensed_medicine_4]):
#             self._add_dispensed_medicine_event(patient_id, visit_id, visit_timestamp, row)
#         if row.presenting_complaint:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Complaint', row.presenting_complaint)
#         if any([row.heart_rate, row.blood_pressure, row.o2_sats,
#                 row.respiratory_rate, row.temperature, row.blood_glucose]):
#             self._add_vitals_event(patient_id, visit_id, visit_timestamp, row)
#         if row.examination:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Examination', row.examination)
#         if row.diagnosis:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Diagnosis', row.diagnosis)
#         if row.treatment:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Treatment', row.treatment)
#         if row.prescription:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Prescriptions', row.prescription)
#         if row.notes:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Notes', row.notes)
#         if row.camp:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Camp', row.camp)

#     def _add_text_event(self, patient_id: str, visit_id: str, visit_timestamp: datetime,
#                         event_type: str, event_metadata: str):
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type=event_type,
#             event_timestamp=visit_timestamp,
#             event_metadata=event_metadata,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_dispensed_medicine_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         data = [
#             (row.dispensed_medicine_1, row.dispensed_medicine_quantity_1),
#             (row.dispensed_medicine_2, row.dispensed_medicine_quantity_2),
#             (row.dispensed_medicine_3, row.dispensed_medicine_quantity_3),
#             (row.dispensed_medicine_4, row.dispensed_medicine_quantity_4),
#         ]
#         content = '\n'.join([': '.join(r) for r in data if all(r)])
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Medicine Dispensed',
#             event_timestamp=visit_timestamp,
#             event_metadata=content,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_vitals_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         try:
#             diastolic, systolic = row.blood_pressure.split('/')
#         except (ValueError, AttributeError):
#             diastolic = None
#             systolic = None

#         data = {
#             'heartRate': as_string(row.heart_rate),
#             'systolic': as_string(systolic),
#             'diastolic': as_string(diastolic),
#             'sats': as_string(row.o2_sats),
#             'temp': as_string(row.temperature),
#             'respiratoryRate': as_string(row.respiratory_rate),
#             'bloodGlucose': as_string(row.blood_glucose)
#         }

#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Vitals',
#             event_timestamp=visit_timestamp,
#             event_metadata=json.dumps(data),
#             edited_at=datetime.now(),
#         )
#         add_event(event)
